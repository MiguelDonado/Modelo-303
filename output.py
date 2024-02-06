from openpyxl import load_workbook
import csv


def write_to_xlsx(meses):
    # The argument is a list with 12 tuples. Each tuples contains the desired data for each month.
    # It creates a xlsx with the desired format.

    fieldnames = get_cabeceras()

    # Load the existing workbook
    wb = load_workbook("MODELO 303.xlsx")

    # Select the first worksheet
    ws = wb.worksheets[0]

    # Get the maximum row containing data
    max_row = ws.max_row

    # Delete all rows in a range
    ws.delete_rows(1, max_row)

    # Write new data starting from A1
    ws.append(fieldnames)
    for mes in meses:
        ws.append(mes)

    # Save the workbook
    wb.save("MODELO 303.xlsx")


def get_cabeceras():
    # This function returns the cabeceras that are used in Dossier Anual HH-2
    with open("cabeceras.csv", "r") as file:
        reader = csv.reader(file)
        cabeceras = [row[0] for row in reader]  # Get the first column from each row
        return cabeceras
